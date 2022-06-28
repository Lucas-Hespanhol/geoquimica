import openpyxl
import numpy as np

def find_chn(formula): # Analisa a formula molecular e retorna os numeros de C, H e N
    i_c = formula.find("C")
    i_h = formula.find("H")
    i_n = formula.find("N")
    i_o = formula.find("O")
    i_s = formula.find("S")

    if i_h - i_c == 1:
        c = 1
    else:
        c = int(formula[i_c + 1 : i_h])
    
    if i_n != -1 and i_n - i_h == 1:
        h = 1
    elif i_n != -1:
        h = int(formula[i_h + 1 : i_n])
    elif i_o != -1 and i_o - i_h == 1:
        h = 1
    elif i_o != -1:
        h = int(formula[i_h + 1 : i_o])
    elif i_s != -1 and i_s - i_h == 1:
        h = 1
    elif i_s != -1:
        h = int(formula[i_h + 1 : i_s])
    else:
        h = int(formula[i_h + 1 :])
    
    if i_n != -1:
        l = min([ i for i in [i_o, i_s, len(formula)] if i != -1 ])
        if l - i_n == 1:
            n = 1
        else:
            n = int(formula[i_n + 1 : l])
    else:
        n = 0

    if i_o != -1:
        l = min([ i for i in [i_s, len(formula)] if i != -1 ])
        if l - i_o == 1:
            o = 1
        else:
            o = int(formula[i_o + 1 : l])
    else:
        o = 0

    if i_s != -1:
        l = min([ i for i in [len(formula)] if i != -1 ])
        if l - i_s == 1:
            s = 1
        else:
            s = int(formula[i_s + 1 : l])
    else:
        s = 0
    
    return c, h, n, o, s

# class C_ComponenteQuimico{ // Armazena os dados relevantes de cada componente
#     public:
#         std::string cls;
#         std::string mol_formula;
#         float intensity;
#         int C;
#         int H;
#         int N;
#         int DBE;
#         C_ComponenteQuimico(std::vector<std::string> input_cols); // Obtem os dados de um vetor com as strings de cada coluna de cada linha da entrada
#         std::string to_line(); // Imprime os dados em uma string para a saida
# };

class Component:
    def __init__(self, hclass, formula, intensity):
        self.hclass = hclass
        self.formula = formula
        self.intensity = intensity
        self.C, self.H, self.N, self.O, self.S = find_chn(formula)
        self.dbe = self.C - (self.H/2) + (self.N/2) + 1
        self.data = { "Class" : hclass, "Formula" : formula, "Intensity": intensity,
                      "C": self.C, "H": self.H, "N": self.N, "O": self.O, "S": self.S,
                      "H/C": self.H/self.C, "N/C": self.N/self.C, "O/C": self.O/self.C, 
                      "S/C": self.O/self.C, "DBE": self.dbe }
                      

# class C_DBE{ // Armazena os dados relevantes de um grupo de componentes com o mesmo C_DBE
#     public:
#         component_vector dbe_components;
#         std::string parent_class = "";
#         int val;
#         float intensity = 0;
#         std::map<int, struct intensity> c_intensity;
#         int min_c = 10000; // Guardara o C do componente com o menor C no C_DBE
#         int max_c = -1; // Guardara o C do componente com o maior C no C_DBE
#         float sum_c(int min, int max, int opt);
#         float sum_c(std::vector<int> vals);
#         C_DBE(component_vector components, int C_DBE);
#         void print_intensity_per_c(std::ostream &output);

#         // A ser usado na variavel opt da funcao sum_c para escolher entre somar qualquer C, os Impares e os Pares
#         static const int all = 0; 
#         static const int odd = 1;
#         static const int even = 2;
# };

class DBE:
    def __init__(self, dbe, parent_class=None):
        self.dbe = dbe
        self.componens = []
        self.intensity = 0
        self.c_intensity = dict()
        self.parent_class = parent_class

    def add_component(self, component):
        if component.dbe == self.dbe:
            self.componens.append(component)
            self.intensity += component.intensity
            if component.C not in self.c_intensity:
                self.c_intensity[component.C] = 0
            self.c_intensity[component.C] += component.intensity

    def sum_c(self, vals, even=False, odd=False):
        result = 0

        if even:
            vals = filter(lambda x : (x % 2 == 0), vals)
        if odd:
            vals = filter(lambda x : (x % 2 == 1), vals)

        for v in vals:
            if v in self.c_intensity:
                result += self.c_intensity[v]

        return result

# class C_Heteroatomica{ // Armazena os dados importantes de uma classe heteroatomica
#     public:
#         component_vector class_components;
#         std::string class_name;
#         float intensity = 0;
#         std::map<int, C_DBE*> class_dbes;
#         float sum_dbe(int min, int max);
#         float sum_dbe(std::vector<int> indices);
#         C_Heteroatomica(component_vector components, std::string name);
#         C_DBE *get_DBE(int C_DBE); // Retorna um ponteiro ao C_DBE especificado. Se não existir, inicializa um objeto novo e retorna
#         float C(int C_DBE, int c);

#         // Chamam a funcao sum_c do C_DBE, simplifica a sintaxe de chamar a funcao do ponteiro dentro do mapa
#         float sum_C(int C_DBE, int min, int max, int opt); 
#         float sum_C(int C_DBE, std::vector<int> vals);

#         void print_intensity_per_dbe(std::ostream &output);
# };

class HClass:
    def __init__(self, name):
        self.name = name
        self.components = []
        self.dbes = dict()
        self.intensity = 0
        pass

    def add_component(self, component):
        if component.hclass == self.name:
            self.components.append(component)
            dbe = component.dbe
            if dbe not in self.dbes:
                self.dbes[dbe] = DBE(dbe, parent_class=self)
            self.dbes[dbe].add_component(component)
            self.intensity += component.intensity

    def sum_dbe(self, vals, even=False, odd=False):
        result = 0

        if even:
            vals = filter(lambda x : (x % 2 == 0), vals)
        if odd:
            vals = filter(lambda x : (x % 2 == 1), vals)

        for v in vals:
            if v in self.dbes.keys():
                result += self.dbes.get(v, DBE(v)).intensity
                
        return result

    def sum_C(self, dbe, s, e, even=False, odd=False):
        vals = range(s, e+1)
        r = 0
        if even:
            vals = filter(lambda x : (x % 2 == 0), vals)
        if odd:
            vals = filter(lambda x : (x % 2 == 1), vals)
        for v in vals:
            r += self.dbe_c(dbe, v)
        return r
    
    def sum_C_list(self, dbe, vals):
        r = 0
        for v in vals:
            r += self.dbe_c(dbe, v)
        return r

    
    def dbe_c(self, dbe, c):
        if dbe in self.dbes:
            if c in self.dbes.get(dbe, DBE(dbe)).c_intensity:
                return self.dbes.get(dbe, DBE(dbe)).c_intensity[c]
        return 0

# class C_Amostra{ // Armazena os dados refinados de uma amostra completa
#     public:
#         float intensity;
#         component_vector C_Amostra_components;
#         std::map<std::string, C_Heteroatomica*> classes;
#         std::map<int, C_DBE*> dbes;
#         std::vector<std::string> class_names;
#         C_Amostra(component_vector components);
#         C_Heteroatomica *get_class(std::string name); // Se não existir, inicializa um objeto novo e retorna
#         C_DBE *get_DBE(int C_DBE); // Se nao existir, inicializa um objeto novo e retorna
#         void print_relative_abundancy(std::ostream &output);
#         void print_biodegradation(std::ostream &output);
#         void print_paleoenvironment(std::ostream &output);
#         void print_maturity(std::ostream &output);
# };

class Sample:
    def __init__(self, name):
        self.name = name
        self.components = []
        pass

    def read_from_xlsx(self, filename):
        workbook = openpyxl.open(filename)
        sheet = workbook.active
        for col in sheet.iter_cols(0, sheet.max_column):
            name = col[0].value
            if name is not None:
                if "Class" in name:
                    self.class_ = []
                    for a in col[1:]:
                        if a.value is not None:
                            self.class_.append(str(a.value))
                        else:
                            self.class_.append("XX")
                elif "DBE" in name:
                    self.dbe_ = []
                    for a in col[1:]:
                        try:
                            self.dbe_.append(int(a.value))
                        except:
                            self.dbe_.append(0)
                elif "Mol" in name:
                    self.formula_ = []
                    for a in col[1:]:
                        if a.value is not None:
                            self.formula_.append(str(a.value))
                        else:
                            self.formula_.append("C0H0N0")
                elif "Intensity" in name:
                    self.int_ = []
                    for a in col[1:]:
                        try:
                            self.int_.append(float(a.value))
                        except:
                            self.int_.append(0)
        self.process()

    def read_from_csv(self, filename):
        pass

    def process(self):
        # Inicialização das HClasses e DBEs
        self.intensity = sum(self.int_)
        self.hclasses = dict()
        self.dbes = dict()

        for i in range(len(self.class_)):
            if self.class_[i] != "XX":
                c = Component(self.class_[i], self.formula_[i], self.int_[i])
                self.components.append(c)
                if c.hclass not in self.hclasses:
                    self.hclasses[c.hclass] = HClass(c.hclass)
                self.hclasses[c.hclass].add_component(c)

                if c.dbe not in self.dbes:
                    self.dbes[c.dbe] = DBE(c.dbe)
                self.dbes[c.dbe].add_component(c)

        # Análises
        self.biodegradation = dict()
        O1 = self.hclasses.get("O", HClass("O"))
        O2 = self.hclasses.get("O2", HClass("O2"))
        self.biodegradation["A/C"] = np.divide((O2.dbes.get(1, DBE(1)).intensity),O2.sum_dbe([2, 3, 4]))
        self.biodegradation["A/C mod."] = np.divide((O2.dbes.get(1, DBE(1)).intensity),O2.sum_dbe([2, 3, 4, 5, 6]))
        self.biodegradation["Indice S/A (%)"] = 100*O2.sum_dbe([1, 2, 3, 4, 5, 6])/O2.intensity
        self.biodegradation["S/A mod. (%)"] = 100*O2.sum_dbe([2, 3, 4, 5, 6])/O2.intensity
        self.biodegradation["Indice MA1"] = np.divide((O1.dbes.get(4, DBE(4)).intensity),(O1.dbes.get(5, DBE(5)).intensity))
        self.biodegradation["Indice MA2"] = np.divide((O1.dbes.get(4, DBE(4)).intensity),(O1.dbes.get(7, DBE(7)).intensity))
        
        self.paleoenvironment = dict()
        NO = self.hclasses.get("NO", HClass("NO"))
        N = self.hclasses.get("N", HClass("N"))

        self.paleoenvironment["Indice Phenol"] = np.divide((O1.dbes.get(4, DBE(4)).intensity),O1.intensity)
        self.paleoenvironment["C27/C28 (DBE4)"] = np.divide((O1.dbe_c(4, 27)),(O1.dbe_c(4, 28)))
        self.paleoenvironment["C27/C28 (DBE5)"] = np.divide((O1.dbe_c(5, 27)),(O1.dbe_c(5, 28)))
        self.paleoenvironment["Par/Impar (FA)"] = np.divide((O2.sum_C(1, 20, 36, even=True)),(O2.sum_C(1, 19, 35, odd=True)))
        self.paleoenvironment["OEP (FA)"] = np.divide((O2.dbe_c(1, 22) + 6.0*O2.dbe_c(1, 28) + O2.dbe_c(1, 26)),(4*O2.dbe_c(1, 23) + 4*O2.dbe_c(1, 25)))
        self.paleoenvironment["TAR (FA) Par"] = np.divide((O2.sum_C_list(1, [24, 26, 28])),(O2.sum_C_list(1, [12, 14, 16])))
        self.paleoenvironment["TAR (FA) Impar"] = np.divide((O2.sum_C_list(1, [27, 29, 31])),(O2.sum_C_list(1, [15, 17, 19])))
        self.paleoenvironment["C36 Indice de Acido Hepanoico"] = np.divide(O2.dbe_c(6, 36),O2.dbe_c(6, 31))
        self.paleoenvironment["Hopano / Acido Estanoico"] = np.divide(O2.dbe_c(6, 31),O2.dbe_c(5, 28))
        self.paleoenvironment["Razao Rocha 1"] = np.divide(O2.sum_dbe([1,2,3,4,5,6,7]),(O2.sum_dbe([1,2,3,4,5,6,7]) + O2.sum_dbe(range(8, 26))))
        self.paleoenvironment["Razao Rocha 2"] = np.divide(O1.sum_dbe([4,5,6,7,8,9,10]),(O1.sum_dbe([4,5,6,7,8,9,10]) + O1.sum_dbe(range(11, 24))))
        self.paleoenvironment["Razao Rocha 3 (%)"] = 100*NO.dbes.get(10, DBE(0)).intensity/NO.intensity
        self.paleoenvironment["Razao Rocha 4 (%)"] = 100*O1.dbes.get(4, DBE(0)).intensity/O1.intensity
        self.paleoenvironment["Razao Rocha 5"] = np.divide(N.sum_dbe([6,7,8,9,10]),(N.sum_dbe([6,7,8,9,10]) + N.sum_dbe(range(10, 28))))
        self.paleoenvironment["Razao Rocha 6"] = np.divide(NO.sum_dbe(range(2, 15)),(NO.sum_dbe(range(2, 15)) + NO.sum_dbe(range(15, 30))))
        self.paleoenvironment["Razao Rocha 7"] = np.divide((O1.sum_dbe(range(4, 13)) + O2.dbes.get(12, DBE(0)).intensity),(O1.sum_dbe([4, 12, 14, 16]) + O2.sum_dbe([12, 13])))
        self.paleoenvironment["Razao Rocha 8"] = np.divide((N.sum_dbe([8, 9]) + NO.sum_dbe([9, 10])),(N.sum_dbe([8, 9, 13, 16]) + NO.sum_dbe([9, 10, 19, 20])))

        self.maturity = dict()
        self.maturity["DBE 12 da Classe N"] = N.dbes.get(12, DBE(12)).intensity
        self.maturity["DBE 9 da Classe N"] = N.dbes.get(9, DBE(9)).intensity
        self.maturity["DBE 15 da Classe N"] = N.dbes.get(15, DBE(15)).intensity
        self.maturity["DBE 5 da Classe O2"] = O2.dbes.get(5, DBE(5)).intensity
        self.maturity["DBE 6 da Classe O2"] = O2.dbes.get(6, DBE(6)).intensity
        self.maturity["DBE 9 da Classe O2"] = O2.dbes.get(9, DBE(9)).intensity
        self.maturity["DBE 11 da Classe O2"] = O2.dbes.get(11, DBE(11)).intensity
        self.maturity["DBE 13 da Classe O2"] = O2.dbes.get(13, DBE(13)).intensity
        self.maturity["DBE9/DBE12 da Classe N"] = N.dbes.get(9, DBE(9)).intensity/N.dbes.get(12, DBE(12)).intensity
        self.maturity["DBE5/DBE9 da Classe O2"] = O2.dbes.get(5, DBE(5)).intensity/O2.dbes.get(9, DBE(9)).intensity

    def diagnostics(self, filename):
        response = self.name + "\nBiodegradação\n"
        for key in self.biodegradation.keys():
            response += "%s: %.5f\n" % (key, self.biodegradation[key])
        response += "\nPaleoambiente\n"
        for key in self.paleoenvironment.keys():
            response += "%s: %.5f\n" % (key, self.paleoenvironment[key])
        
        response += "\Maturidade\n"
        for key in self.maturity.keys():
            response += "%s: %.5f\n" % (key, self.maturity[key])
        
        f = open("output/" + filename, "w")
        f.write(response)

    def diagnostics_xlsx(self, filename):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = self.name
        ws2 = wb.create_sheet("Biodegradação")
        ws3 = wb.create_sheet("Paleoambiente")
        ws4 = wb.create_sheet("Maturidade")
        ws5 = wb.create_sheet("Distribuição Heteroatômica")

        headers = ["Class", "Formula", "C", "H", "N", "O", "S", "H/C", "N/C", "O/C", "S/C", "DBE", "Intensity"]
        for col in range(1, 1+len(headers)):
            ws1.cell(1, col, value=headers[col-1])
        for row in range(len(self.components)):
            for col in range(1, 1+ len(headers)):
                ws1.cell(2+row, col, value=self.components[row].data[headers[col-1]])

        i = 1
        for k in self.biodegradation.keys():
            ws2.cell(i, 1, value=k)
            ws2.cell(i, 2, value=self.biodegradation[k])
            i += 1

        i = 1
        for k in self.paleoenvironment.keys():
            ws3.cell(i, 1, value=k)
            ws3.cell(i, 2, value=self.paleoenvironment[k])
            i += 1

        i = 1
        for k in self.maturity.keys():
            ws4.cell(i, 1, value=k)
            ws4.cell(i, 2, value=self.maturity[k])
            i += 1


        ws5.cell(1, 1, value="Classe")
        ws5.cell(1, 2, value="Intensidade")
        ws5.cell(1, 3, value="Porcentagem")
        i = 2
        for k in self.hclasses.keys():
            ws5.cell(i, 1, value=k)
            ws5.cell(i, 2, value=self.hclasses[k].intensity)
            ws5.cell(i, 3, value=self.hclasses[k].intensity/self.intensity)
            i += 1

        for k in self.hclasses.keys():
            hc = self.hclasses[k]
            ws = wb.create_sheet(k)
            cs = list()
            for dbe in hc.dbes.keys():
                for c in hc.dbes[dbe].c_intensity.keys():
                    if c not in cs:
                        cs.append(c)
            cs.sort()
            j = 2
            for c in cs:
                ws.cell(1,  j, value=("C%d" % c))
                j += 1
            
            ws.cell(1,  j, value="Total do DBE")

            i = 2
            for dbe in hc.dbes.keys():
                ws.cell(i,  1, value=("DBE%.0f" % dbe))
                j = 2
                for c in cs:
                    if c in hc.dbes[dbe].c_intensity:
                        ws.cell(i, j, value=hc.dbes[dbe].c_intensity[c]/hc.intensity)
                    else:
                        ws.cell(i, j, value=0.0)
                    j += 1
                ws.cell(i, j, value=hc.dbes[dbe].intensity/hc.intensity)
                i += 1
            

        wb.save("output/" + filename)

    def class_relative_abundancy(self, filename):
        result = self.name + "\n"
        hclasses = list(self.hclasses.keys())
        intensities = [ self.hclasses[a].intensity for a in hclasses ]
        intensities = np.divide(intensities, self.intensity)

        for i in range(len(hclasses)):
            result += "%s: %.5f\n" % (hclasses[i], intensities[i])

        f = open("output/" + filename, "w")
        f.write(result)

    def class_dbe(self, cls_):
        hclass = self.hclasses.get(cls_, HClass(cls_))
        x = []
        y = []
        for dbe in sorted(hclass.dbes.keys()):
            x.append(int(dbe))
            y.append(hclass.dbes[dbe].intensity) 
        return x, y